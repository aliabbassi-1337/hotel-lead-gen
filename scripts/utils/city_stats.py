#!/usr/bin/env python3
"""
Sadie City Stats - Generate per-city stats from split files.
Creates a summary showing leads per city.

Usage:
    python3 sadie_city_stats.py detector_output/florida/FL
    python3 sadie_city_stats.py detector_output/florida/FL --output onedrive_output/florida_cities.xlsx
"""

import csv
import os
import sys
import argparse
from datetime import datetime
from collections import Counter

def count_with_field(rows: list, field: str) -> int:
    """Count rows that have a non-empty value for a field."""
    return sum(1 for r in rows if r.get(field, "").strip())

def load_csv(filepath: str) -> list:
    """Load CSV file and return list of rows."""
    if not os.path.exists(filepath):
        return []
    with open(filepath, "r", encoding="utf-8") as f:
        return list(csv.DictReader(f))

def generate_city_stats(city_dir: str) -> list:
    """Generate stats for all cities in a state directory."""
    if not os.path.isdir(city_dir):
        print(f"Error: {city_dir} is not a directory")
        return []

    city_stats = []

    for filename in sorted(os.listdir(city_dir)):
        if not filename.endswith(".csv"):
            continue

        filepath = os.path.join(city_dir, filename)
        rows = load_csv(filepath)

        if not rows:
            continue

        city_name = filename.replace(".csv", "").replace("_", " ").title()

        # Calculate stats
        total = len(rows)
        with_booking_url = count_with_field(rows, "booking_url")
        with_known_engine = sum(1 for r in rows
            if r.get("booking_engine", "").strip() and
            r.get("booking_engine") not in ["unknown", "unknown_third_party", "proprietary_or_same_domain", "contact_only", "", "unknown_booking_api"])
        with_email = count_with_field(rows, "email")
        with_phone = sum(1 for r in rows
            if r.get("phone_google", "").strip() or r.get("phone_website", "").strip())
        with_room_count = sum(1 for r in rows
            if r.get("room_count", "").strip() and r.get("room_count") not in ["", "None", "null"])

        # Engine breakdown
        engine_counts = Counter(r.get("booking_engine", "") for r in rows if r.get("booking_engine", "").strip())
        top_engine = engine_counts.most_common(1)[0][0] if engine_counts else ""

        city_stats.append({
            "city": city_name,
            "total_leads": total,
            "with_booking_url": with_booking_url,
            "booking_rate": round(with_booking_url / total * 100, 1) if total > 0 else 0,
            "with_known_engine": with_known_engine,
            "with_email": with_email,
            "with_phone": with_phone,
            "with_room_count": with_room_count,
            "top_engine": top_engine,
        })

    # Sort by total leads descending
    city_stats.sort(key=lambda x: -x["total_leads"])

    return city_stats

def save_csv(stats: list, output_path: str):
    """Save stats to CSV."""
    if not stats:
        return

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    fieldnames = ["city", "total_leads", "with_booking_url", "booking_rate",
                  "with_known_engine", "with_email", "with_phone", "with_room_count", "top_engine"]

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(stats)

    print(f"Saved to: {output_path}")

def save_excel(stats: list, output_path: str, state_name: str = "Florida"):
    """Save stats to Excel with formatting."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed, saving as CSV instead")
        save_csv(stats, output_path.replace(".xlsx", ".csv"))
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{state_name} Cities"

    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ["City", "Total Leads", "With Booking URL", "Booking Rate %",
               "Known Engine", "With Email", "With Phone", "Room Count", "Top Engine"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data
    for row_idx, city in enumerate(stats, 2):
        ws.cell(row=row_idx, column=1, value=city["city"]).border = thin_border
        ws.cell(row=row_idx, column=2, value=city["total_leads"]).border = thin_border
        ws.cell(row=row_idx, column=3, value=city["with_booking_url"]).border = thin_border
        ws.cell(row=row_idx, column=4, value=city["booking_rate"]).border = thin_border
        ws.cell(row=row_idx, column=5, value=city["with_known_engine"]).border = thin_border
        ws.cell(row=row_idx, column=6, value=city["with_email"]).border = thin_border
        ws.cell(row=row_idx, column=7, value=city["with_phone"]).border = thin_border
        ws.cell(row=row_idx, column=8, value=city["with_room_count"]).border = thin_border
        ws.cell(row=row_idx, column=9, value=city["top_engine"]).border = thin_border

    # Totals row
    total_row = len(stats) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=sum(s["total_leads"] for s in stats)).font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=sum(s["with_booking_url"] for s in stats)).font = Font(bold=True)
    total_leads = sum(s["total_leads"] for s in stats)
    total_booking = sum(s["with_booking_url"] for s in stats)
    ws.cell(row=total_row, column=4, value=round(total_booking / total_leads * 100, 1) if total_leads > 0 else 0).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=sum(s["with_known_engine"] for s in stats)).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=sum(s["with_email"] for s in stats)).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=sum(s["with_phone"] for s in stats)).font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=sum(s["with_room_count"] for s in stats)).font = Font(bold=True)

    # Column widths
    col_widths = [25, 12, 18, 14, 14, 12, 12, 12, 25]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Save
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"Saved to: {output_path}")

def print_summary(stats: list, state_name: str = "Florida"):
    """Print a summary table to console."""
    if not stats:
        print("No stats to display")
        return

    total_leads = sum(s["total_leads"] for s in stats)
    total_booking = sum(s["with_booking_url"] for s in stats)

    print(f"\n{'='*70}")
    print(f"{state_name.upper()} CITY BREAKDOWN - {len(stats)} Cities, {total_leads} Total Leads")
    print(f"{'='*70}")
    print(f"{'City':<30} {'Leads':>8} {'Booking':>8} {'Rate':>8} {'Engine':>8}")
    print("-" * 70)

    for city in stats[:25]:  # Top 25
        print(f"{city['city']:<30} {city['total_leads']:>8} {city['with_booking_url']:>8} {city['booking_rate']:>7.1f}% {city['with_known_engine']:>8}")

    if len(stats) > 25:
        print(f"... and {len(stats) - 25} more cities")

    print("-" * 70)
    print(f"{'TOTAL':<30} {total_leads:>8} {total_booking:>8} {round(total_booking/total_leads*100, 1) if total_leads > 0 else 0:>7.1f}%")
    print(f"{'='*70}\n")

def main():
    parser = argparse.ArgumentParser(description="Generate per-city stats from split files")
    parser.add_argument("city_dir", help="Directory containing city CSV files (e.g., detector_output/florida/FL)")
    parser.add_argument("--output", "-o", help="Output file path (.csv or .xlsx)")
    parser.add_argument("--state", default="Florida", help="State name for the report title")

    args = parser.parse_args()

    # Generate stats
    stats = generate_city_stats(args.city_dir)

    if not stats:
        print("No city data found")
        return

    # Print summary
    print_summary(stats, args.state)

    # Save output
    if args.output:
        if args.output.endswith(".xlsx"):
            save_excel(stats, args.output, args.state)
        else:
            save_csv(stats, args.output)
    else:
        # Auto-generate output
        parent_dir = os.path.dirname(args.city_dir)
        state_slug = os.path.basename(args.city_dir).lower()
        output_csv = os.path.join(parent_dir, f"{state_slug}_city_stats.csv")
        save_csv(stats, output_csv)

if __name__ == "__main__":
    main()
