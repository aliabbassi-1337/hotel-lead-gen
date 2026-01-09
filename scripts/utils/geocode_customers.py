#!/usr/bin/env python3
"""
Geocode Existing Customers
===========================
Adds lat/lng coordinates to existing customers by searching Google Maps via Serper.

Usage:
    python3 sadie_geocode_customers.py --input data/existing_customers.xlsx
    python3 sadie_geocode_customers.py --input data/existing_customers.xlsx --output data/existing_customers_geocoded.xlsx

Requires:
    - SERPER_API_KEY in .env file
    - pip3 install openpyxl httpx python-dotenv
"""

import os
import re
import asyncio
import argparse
from datetime import datetime
from typing import Optional, Dict

import httpx
from dotenv import load_dotenv

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    print("Error: openpyxl not installed. Run: pip3 install openpyxl")
    exit(1)

load_dotenv()

# Use dedicated Serper key for geocoding
SERPER_API_KEY = os.getenv("SERPER_SAMI") or os.getenv("SERPER_API_KEY")
SERPER_URL = "https://google.serper.dev/places"


def log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")


async def search_hotel_location(client: httpx.AsyncClient, hotel_name: str, location: str) -> Optional[Dict]:
    """Search for a hotel on Google Maps and return its coordinates."""
    
    # Build search query
    query = f"{hotel_name} hotel {location}"
    
    try:
        resp = await client.post(
            SERPER_URL,
            headers={
                "X-API-KEY": SERPER_API_KEY,
                "Content-Type": "application/json",
            },
            json={"q": query, "num": 1},
            timeout=15.0,
        )
        
        if resp.status_code != 200:
            return None
        
        data = resp.json()
        places = data.get("places", [])
        
        if places:
            place = places[0]
            return {
                "lat": place.get("latitude"),
                "lon": place.get("longitude"),
                "address": place.get("address", ""),
                "phone": place.get("phoneNumber", ""),
            }
        
        return None
        
    except Exception as e:
        log(f"  Error searching {hotel_name}: {e}")
        return None


async def geocode_customers(input_file: str, output_file: str):
    """Geocode all customers in the Excel file."""
    
    if not SERPER_API_KEY:
        print("❌ Error: SERPER_API_KEY not found in .env")
        return
    
    if not os.path.exists(input_file):
        print(f"❌ File not found: {input_file}")
        return
    
    # Load Excel
    wb = load_workbook(input_file)
    ws = wb.active
    
    # Get headers
    headers = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
    
    # Find columns
    def find_col(keywords):
        for i, h in enumerate(headers):
            if any(kw in h for kw in keywords):
                return i
        return None
    
    hotel_col = find_col(["hotel", "name", "property"])
    location_col = find_col(["location", "city", "state", "region"])
    email_col = find_col(["email"])
    
    if hotel_col is None:
        log("❌ Could not find 'Hotel' column")
        return
    
    # Add lat/lng columns if they don't exist
    lat_col = find_col(["lat", "latitude"])
    lon_col = find_col(["lon", "lng", "longitude"])
    
    if lat_col is None:
        lat_col = len(headers)
        ws.cell(row=1, column=lat_col + 1, value="Latitude")
        headers.append("latitude")
    
    if lon_col is None:
        lon_col = len(headers)
        ws.cell(row=1, column=lon_col + 1, value="Longitude")
        headers.append("longitude")
    
    # Count rows
    total_rows = ws.max_row - 1  # Exclude header
    log(f"Found {total_rows} customers to geocode")
    
    # Geocode each customer
    geocoded_count = 0
    skipped_count = 0
    
    async with httpx.AsyncClient() as client:
        for row_idx in range(2, ws.max_row + 1):
            hotel_name = ws.cell(row=row_idx, column=hotel_col + 1).value
            location = ws.cell(row=row_idx, column=location_col + 1).value if location_col is not None else ""
            
            # Check if already has coordinates
            existing_lat = ws.cell(row=row_idx, column=lat_col + 1).value
            existing_lon = ws.cell(row=row_idx, column=lon_col + 1).value
            
            if existing_lat and existing_lon:
                skipped_count += 1
                continue
            
            if not hotel_name:
                continue
            
            log(f"Geocoding: {hotel_name}")
            
            # Search for the hotel
            result = await search_hotel_location(client, str(hotel_name), str(location or ""))
            
            if result and result.get("lat") and result.get("lon"):
                ws.cell(row=row_idx, column=lat_col + 1, value=result["lat"])
                ws.cell(row=row_idx, column=lon_col + 1, value=result["lon"])
                geocoded_count += 1
                log(f"  ✓ Found: {result['lat']}, {result['lon']}")
            else:
                log(f"  ✗ Not found")
            
            # Rate limit - Serper is generous but let's be safe
            await asyncio.sleep(0.5)
            
            # Progress
            if (row_idx - 1) % 20 == 0:
                log(f"Progress: {row_idx - 1}/{total_rows}")
    
    # Save
    wb.save(output_file)
    
    log(f"\n✅ Geocoding complete")
    log(f"   Geocoded: {geocoded_count}")
    log(f"   Skipped (already had coords): {skipped_count}")
    log(f"   Output: {output_file}")


def main():
    parser = argparse.ArgumentParser(description="Geocode existing customers")
    parser.add_argument("--input", required=True, help="Input Excel file")
    parser.add_argument("--output", help="Output file (default: overwrites input)")
    
    args = parser.parse_args()
    
    output = args.output or args.input
    
    asyncio.run(geocode_customers(args.input, output))


if __name__ == "__main__":
    main()

