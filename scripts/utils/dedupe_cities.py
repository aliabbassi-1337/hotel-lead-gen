#!/usr/bin/env python3
"""
Dedupe hotels across city Excel files.
Keeps each hotel only in the city nearest to its coordinates.

Usage:
    python3 scripts/utils/dedupe_cities.py --state florida
    python3 scripts/utils/dedupe_cities.py --state florida --dry-run
"""

import os
import sys
import argparse
import math
from pathlib import Path
from collections import defaultdict
from datetime import datetime

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl required. pip install openpyxl")
    sys.exit(1)

ONEDRIVE_BASE = os.path.expanduser("~/Library/CloudStorage/OneDrive-ValsoftCorporation/Sadie Lead Gen")

# City centers for distance calculation
CITY_CENTERS = {
    # Florida
    'miami': (25.76, -80.19),
    'miami beach': (25.79, -80.13),
    'fort lauderdale': (26.12, -80.14),
    'hollywood': (26.01, -80.15),
    'pompano beach': (26.24, -80.13),
    'boca raton': (26.36, -80.08),
    'west palm beach': (26.71, -80.05),
    'palm beach': (26.71, -80.04),
    'delray beach': (26.46, -80.07),
    'key west': (24.56, -81.78),
    'key largo': (25.09, -80.45),
    'islamorada': (24.92, -80.63),
    'marathon': (24.71, -81.09),
    'naples': (26.14, -81.79),
    'marco island': (25.94, -81.72),
    'fort myers': (26.64, -81.87),
    'fort myers beach': (26.45, -81.95),
    'cape coral': (26.56, -81.95),
    'sanibel': (26.44, -82.10),
    'captiva': (26.53, -82.19),
    'sarasota': (27.34, -82.53),
    'siesta key': (27.27, -82.55),
    'bradenton': (27.50, -82.57),
    'bradenton beach': (27.47, -82.70),
    'anna maria island': (27.53, -82.73),
    'longboat key': (27.41, -82.66),
    'clearwater': (27.97, -82.80),
    'clearwater beach': (27.98, -82.83),
    'st petersburg': (27.77, -82.64),
    'treasure island': (27.77, -82.77),
    'madeira beach': (27.80, -82.80),
    'tampa': (27.95, -82.46),
    'orlando': (28.54, -81.38),
    'kissimmee': (28.29, -81.41),
    'celebration': (28.32, -81.54),
    'lake buena vista': (28.37, -81.52),
    'daytona beach': (29.21, -81.02),
    'st augustine': (29.90, -81.31),
    'jacksonville': (30.33, -81.66),
    'jacksonville beach': (30.29, -81.39),
    'amelia island': (30.67, -81.44),
    'fernandina beach': (30.67, -81.44),
    'pensacola': (30.42, -87.22),
    'pensacola beach': (30.33, -87.14),
    'destin': (30.39, -86.50),
    'fort walton beach': (30.42, -86.62),
    'panama city beach': (30.18, -85.80),
    'tallahassee': (30.44, -84.28),
    'gainesville': (29.65, -82.32),
    'cocoa beach': (28.32, -80.61),
    'melbourne': (28.08, -80.61),
    'vero beach': (27.64, -80.40),
    'stuart': (27.20, -80.25),
    'jupiter': (26.93, -80.09),
    'palm coast': (29.58, -81.21),
    'flagler beach': (29.47, -81.13),
    'new smyrna beach': (29.03, -80.93),
    'crystal river': (28.90, -82.59),
    'homosassa': (28.78, -82.62),
    'tarpon springs': (28.15, -82.76),
    'dunedin': (28.02, -82.77),
    'winter park': (28.60, -81.34),
    'apalachicola': (29.73, -84.98),
    'port st joe': (29.81, -85.30),
    'sebring': (27.50, -81.44),
    'punta gorda': (26.93, -82.05),
    'port charlotte': (26.97, -82.09),
    'englewood': (26.96, -82.35),
    'venice': (27.10, -82.45),
    'port st lucie': (27.29, -80.35),
    'hallandale': (25.98, -80.15),
    'aventura': (25.96, -80.14),
    'sunny isles': (25.95, -80.12),
    'dania beach': (26.05, -80.14),
    'lauderdale by the sea': (26.19, -80.10),
    'brickell miami': (25.76, -80.19),
    'indian rocks beach': (27.88, -82.85),
    'florida city': (25.45, -80.48),
    'ormond beach': (29.29, -81.06),
}


def haversine_km(lat1, lon1, lat2, lon2):
    """Calculate distance between two points in km."""
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    return R * 2 * math.asin(math.sqrt(a))


def get_city_center(city_name: str) -> tuple:
    """Get city center coordinates."""
    name_lower = city_name.lower().strip()
    return CITY_CENTERS.get(name_lower)


def normalize_hotel_key(name: str, website: str) -> str:
    """Create a normalized key for deduplication."""
    name = (name or "").strip().lower()
    # Remove common suffixes
    for suffix in [' hotel', ' motel', ' inn', ' resort', ' suites', ' lodge']:
        name = name.replace(suffix, '')
    return name


def extract_city_from_address(address: str) -> str:
    """Extract city name from address string."""
    if not address:
        return None

    import re
    address = str(address)

    # Pattern: "City, FL" or "City, FL 33XXX"
    patterns = [
        r',\s*([A-Za-z\s\-]+),\s*FL\s*,?\s*\d{5}',
        r',\s*([A-Za-z\s\-]+),\s*FL\b',
        r',\s*([A-Za-z\s\-]+),\s*Florida\b',
        r',\s*([A-Za-z\s\-]+)\s+FL\s+\d{5}',
    ]

    for pattern in patterns:
        match = re.search(pattern, address, re.IGNORECASE)
        if match:
            city = match.group(1).strip()
            if city and len(city) > 1 and city.lower() not in ('fl', 'florida'):
                return city.lower()

    return None


def read_excel_leads(xlsx_path: Path) -> list[dict]:
    """Read leads from Excel file."""
    leads = []
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        if "Leads" not in wb.sheetnames:
            wb.close()
            return leads

        ws = wb["Leads"]
        headers = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                headers = [str(h or "").strip() for h in row]
                continue

            lead = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    lead[headers[i]] = val

            if lead.get("name"):
                leads.append(lead)

        wb.close()
    except Exception as e:
        print(f"  Error reading {xlsx_path.name}: {e}")

    return leads


def write_excel_leads(xlsx_path: Path, leads: list[dict], city_name: str):
    """Write leads to Excel file with Stats sheet."""
    if not leads:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Headers
    headers = list(leads[0].keys())
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Data
    for row_idx, lead in enumerate(leads, 2):
        for col, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col, value=lead.get(h))

    # Simple Stats sheet
    ws_stats = wb.create_sheet("Stats")
    ws_stats.cell(row=1, column=1, value=f"{city_name} - {len(leads)} leads")
    ws_stats.cell(row=2, column=1, value=f"Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    wb.save(xlsx_path)


def normalize_city_name(city: str) -> str:
    """Normalize city name for matching."""
    if not city:
        return None
    city = city.lower().strip()
    # Common variations
    city = city.replace("lauderdale-by-the-sea", "lauderdale by the sea")
    city = city.replace("ft ", "fort ")
    city = city.replace("ft. ", "fort ")
    city = city.replace("st ", "st ")
    city = city.replace("st. ", "st ")
    return city


def dedupe_state(state_path: Path, dry_run: bool = False) -> dict:
    """Dedupe hotels across all city files in a state."""

    print(f"\n📁 Processing: {state_path}")

    # Step 1: Read all hotels from all city files
    all_hotels = {}  # key -> (hotel_data, best_city, is_address_match)
    city_files = {}  # city_name -> xlsx_path
    city_name_variants = {}  # normalized_name -> original_city_name

    for xlsx in sorted(state_path.glob("*.xlsx")):
        if xlsx.name.startswith("~") or "Stats" in xlsx.name:
            continue

        city_name = xlsx.stem
        city_files[city_name] = xlsx
        city_name_variants[normalize_city_name(city_name)] = city_name

        leads = read_excel_leads(xlsx)
        print(f"  {city_name}: {len(leads)} leads")

        for lead in leads:
            key = normalize_hotel_key(lead.get("name", ""), lead.get("website", ""))
            if not key:
                continue

            # Try to extract city from address
            address = lead.get("address", "")
            address_city = extract_city_from_address(address)
            address_city_normalized = normalize_city_name(address_city) if address_city else None

            # Check if address city matches current file's city
            current_city_normalized = normalize_city_name(city_name)
            is_address_match = address_city_normalized == current_city_normalized if address_city_normalized else False

            # If address explicitly says a different city, use that city
            # BUT only if that city already has a file (don't create new cities)
            best_city = city_name
            if address_city_normalized and address_city_normalized in city_name_variants:
                best_city = city_name_variants[address_city_normalized]
                is_address_match = True

            # Keep hotel - prefer address match, otherwise first occurrence
            if key in all_hotels:
                existing = all_hotels[key]
                # Only replace if new one has address match and existing doesn't
                if is_address_match and not existing[2]:
                    all_hotels[key] = (lead, best_city, is_address_match)
            else:
                all_hotels[key] = (lead, best_city, is_address_match)

    # Step 2: Group hotels by their best city
    city_leads = defaultdict(list)
    for key, (lead, city_name, dist) in all_hotels.items():
        city_leads[city_name].append(lead)

    # Step 3: Calculate stats
    total_before = sum(len(read_excel_leads(xlsx)) for xlsx in city_files.values())
    total_after = sum(len(leads) for leads in city_leads.values())
    duplicates_removed = total_before - total_after

    print(f"\n📊 Deduplication results:")
    print(f"   Before: {total_before} leads across {len(city_files)} cities")
    print(f"   After:  {total_after} unique leads")
    print(f"   Duplicates removed: {duplicates_removed} ({duplicates_removed/total_before*100:.1f}%)")

    # Step 4: Write updated files (never delete, never create new)
    if not dry_run:
        print(f"\n💾 Updating files...")

        for city_name, xlsx_path in city_files.items():
            leads = city_leads.get(city_name, [])
            before_count = len(read_excel_leads(xlsx_path))

            if leads:
                write_excel_leads(xlsx_path, leads, city_name)
                if len(leads) != before_count:
                    print(f"   ✓ {city_name}: {before_count} → {len(leads)} leads")
                else:
                    print(f"   · {city_name}: {len(leads)} leads (unchanged)")
            else:
                # Keep empty file as-is, don't delete
                print(f"   · {city_name}: 0 leads (kept)")
    else:
        print(f"\n🔍 Dry run - no files modified")
        print(f"\nChanges per city:")
        for city_name, xlsx_path in sorted(city_files.items()):
            before = len(read_excel_leads(xlsx_path))
            after = len(city_leads.get(city_name, []))
            if before != after:
                diff = before - after
                direction = "removed" if diff > 0 else "added"
                print(f"   {city_name}: {before} → {after} ({abs(diff)} {direction})")

    return {"before": total_before, "after": total_after, "removed": duplicates_removed}


def main():
    parser = argparse.ArgumentParser(description="Dedupe hotels across city files")
    parser.add_argument("--country", default="USA", help="Country folder (default: USA)")
    parser.add_argument("--state", required=True, help="State to dedupe (e.g., florida)")
    parser.add_argument("--dry-run", action="store_true", help="Show what would change without modifying files")
    args = parser.parse_args()

    state_path = Path(ONEDRIVE_BASE) / args.country / args.state.title()

    if not state_path.exists():
        print(f"❌ State folder not found: {state_path}")
        sys.exit(1)

    stats = dedupe_state(state_path, dry_run=args.dry_run)

    print(f"\n✅ Done!")


if __name__ == "__main__":
    main()
