#!/usr/bin/env python3
"""
Sadie Existing Customer Enricher
=================================
Adds "Existing Customer in Area" field to leads by finding nearby Sadie hotels.

Usage:
    python3 sadie_existing_customer_enricher.py --input detector_output/sydney_leads_post.csv
    python3 sadie_existing_customer_enricher.py --input detector_output/sydney_leads_post.csv --customers data/existing_customers.xlsx

Requires:
    - data/existing_customers.xlsx (download from SharePoint)
    - pip3 install openpyxl
"""

import os
import csv
import math
import argparse
from datetime import datetime
from typing import Optional, List, Dict

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed. Run: pip3 install openpyxl")
    exit(1)


DEFAULT_CUSTOMERS_FILE = "data/existing_customers.xlsx"


def log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")


def haversine_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calculate distance between two points in kilometers using Haversine formula."""
    R = 6371  # Earth's radius in km
    
    lat1_rad = math.radians(lat1)
    lat2_rad = math.radians(lat2)
    delta_lat = math.radians(lat2 - lat1)
    delta_lon = math.radians(lon2 - lon1)
    
    a = math.sin(delta_lat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(delta_lon/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    
    return R * c


def load_existing_customers(filepath: str) -> List[Dict]:
    """Load existing Sadie customers from Excel file."""
    if not os.path.exists(filepath):
        log(f"❌ Customers file not found: {filepath}")
        log(f"   Download from SharePoint and save to: {filepath}")
        return []
    
    customers = []
    
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).lower().strip() if cell.value else "")
        
        # Find relevant columns (flexible matching)
        def find_col_exact(keyword):
            """Find column with exact match."""
            for i, h in enumerate(headers):
                if h == keyword:
                    return i
            return None
        
        def find_col(keywords):
            """Find column containing any keyword."""
            for i, h in enumerate(headers):
                if any(kw in h for kw in keywords):
                    return i
            return None
        
        # Use exact matches where possible
        name_col = find_col_exact("hotel")
        lat_col = find_col(["latitude"])
        lon_col = find_col(["longitude"])
        gm_col = find_col_exact("name")  # GM/contact name
        phone_col = find_col_exact("phone number")
        city_col = find_col(["location"])
        
        if name_col is None:
            log(f"⚠ Could not find 'name' column in {filepath}")
            log(f"   Headers found: {headers}")
            return []
        
        # Read data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[name_col]:
                continue
            
            customer = {
                "name": str(row[name_col]).strip() if row[name_col] else "",
                "lat": None,
                "lon": None,
                "gm": "",
                "phone": "",
                "city": "",
            }
            
            # Extract lat/lon
            if lat_col is not None and row[lat_col]:
                try:
                    customer["lat"] = float(row[lat_col])
                except:
                    pass
            
            if lon_col is not None and row[lon_col]:
                try:
                    customer["lon"] = float(row[lon_col])
                except:
                    pass
            
            # Extract optional fields
            if gm_col is not None and len(row) > gm_col and row[gm_col]:
                customer["gm"] = str(row[gm_col]).strip()
            
            if phone_col is not None and len(row) > phone_col and row[phone_col]:
                customer["phone"] = str(row[phone_col]).strip()
            
            if city_col is not None and len(row) > city_col and row[city_col]:
                customer["city"] = str(row[city_col]).strip()
            
            if customer["name"]:
                customers.append(customer)
        
        wb.close()
        
    except Exception as e:
        log(f"❌ Error reading customers file: {e}")
        return []
    
    log(f"✓ Loaded {len(customers)} existing Sadie customers")
    
    # Count how many have coordinates
    with_coords = sum(1 for c in customers if c["lat"] and c["lon"])
    log(f"  {with_coords} have lat/lng coordinates")
    
    return customers


def find_nearest_customer(lead_lat: float, lead_lon: float, customers: List[Dict], max_distance_km: float = 100) -> Optional[Dict]:
    """Find the nearest existing customer within max_distance_km."""
    nearest = None
    min_distance = float('inf')
    
    for customer in customers:
        if not customer["lat"] or not customer["lon"]:
            continue
        
        distance = haversine_distance(lead_lat, lead_lon, customer["lat"], customer["lon"])
        
        if distance < min_distance and distance <= max_distance_km:
            min_distance = distance
            nearest = {**customer, "distance_km": round(distance, 1)}
    
    return nearest


def format_customer_info(customer: Dict) -> str:
    """Format customer info for the 'Existing Customer in Area' field."""
    if not customer:
        return ""
    
    parts = []
    
    # Name and distance
    name_part = f"Nearest: {customer['name']} ({customer['distance_km']}km)"
    parts.append(name_part)
    
    # GM if available
    if customer.get("gm"):
        parts.append(f"GM: {customer['gm']}")
    
    # Phone if available
    if customer.get("phone"):
        parts.append(f"Phone: {customer['phone']}")
    
    return " | ".join(parts)


def enrich_leads(input_file: str, output_file: str, customers_file: str, max_distance_km: float = 100):
    """Enrich leads with nearest existing customer info."""
    
    # Load existing customers
    customers = load_existing_customers(customers_file)
    
    if not customers:
        log("❌ No customers loaded. Cannot enrich leads.")
        return
    
    # Load leads
    with open(input_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        leads = list(reader)
        fieldnames = list(reader.fieldnames)
    
    if not leads:
        log(f"❌ No leads in {input_file}")
        return
    
    log(f"✓ Loaded {len(leads)} leads from {input_file}")
    
    # Ensure "Existing Customer in Area" column exists
    if "Existing Customer in Area" not in fieldnames:
        fieldnames.append("Existing Customer in Area")
    
    # Find lat/lon columns in leads
    lat_col = None
    lon_col = None
    for col in fieldnames:
        if col.lower() in ["lat", "latitude"]:
            lat_col = col
        elif col.lower() in ["lon", "lng", "longitude"]:
            lon_col = col
    
    if not lat_col or not lon_col:
        log("⚠ Leads file doesn't have lat/lng columns. Trying 'latitude'/'longitude'...")
        lat_col = "latitude"
        lon_col = "longitude"
    
    # Enrich each lead
    enriched_count = 0
    
    for lead in leads:
        # Skip if already has value
        if lead.get("Existing Customer in Area", "").strip():
            continue
        
        # Get lead coordinates
        try:
            lead_lat = float(lead.get(lat_col, ""))
            lead_lon = float(lead.get(lon_col, ""))
        except (ValueError, TypeError):
            lead["Existing Customer in Area"] = ""
            continue
        
        # Find nearest customer
        nearest = find_nearest_customer(lead_lat, lead_lon, customers, max_distance_km)
        
        if nearest:
            lead["Existing Customer in Area"] = format_customer_info(nearest)
            enriched_count += 1
        else:
            lead["Existing Customer in Area"] = ""
    
    # Save enriched leads
    with open(output_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(leads)
    
    log(f"✅ Enriched {enriched_count} leads with nearby existing customers")
    log(f"   Output: {output_file}")


def main():
    parser = argparse.ArgumentParser(description="Add 'Existing Customer in Area' to leads")
    parser.add_argument("--input", required=True, help="Input leads CSV file")
    parser.add_argument("--output", help="Output CSV file (default: overwrites input)")
    parser.add_argument("--customers", default=DEFAULT_CUSTOMERS_FILE, help="Existing customers Excel file")
    parser.add_argument("--max-distance", type=float, default=100, help="Max distance in km (default: 100)")
    
    args = parser.parse_args()
    
    output_file = args.output or args.input
    
    enrich_leads(
        input_file=args.input,
        output_file=output_file,
        customers_file=args.customers,
        max_distance_km=args.max_distance
    )


if __name__ == "__main__":
    main()

