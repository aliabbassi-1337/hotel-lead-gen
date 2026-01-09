#!/usr/bin/env python3
"""
Sadie Country Stats - Generate Aggregate Stats per Country
===========================================================
Scans OneDrive folder structure and creates country-level aggregate stats.

Usage:
    python3 sadie_country_stats.py
    python3 sadie_country_stats.py --country usa
"""

import os
import sys
import argparse
from pathlib import Path
from datetime import datetime
from collections import defaultdict

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl not installed. Install with: pip3 install openpyxl")
    sys.exit(1)

# OneDrive paths
ONEDRIVE_BASE = os.path.expanduser("~/Library/CloudStorage/OneDrive-ValsoftCorporation")
SADIE_FOLDER = "Sadie Lead Gen"


def get_sadie_path() -> Path:
    return Path(ONEDRIVE_BASE) / SADIE_FOLDER


def find_city_files(country_path: Path) -> list[Path]:
    """Find the most recent Excel file per city folder (recursively)."""
    # Group files by their parent directory (city)
    city_files = defaultdict(list)
    
    for xlsx in country_path.rglob("*.xlsx"):
        # Skip country stats files themselves
        if "_country_stats" in xlsx.name:
            continue
        # Skip temp files
        if xlsx.name.startswith("~"):
            continue
        # Skip files in the country root (not in a city folder)
        if xlsx.parent == country_path:
            continue
            
        city_files[xlsx.parent].append(xlsx)
    
    # For each city, keep only the most recent file (by modification time)
    latest_files = []
    for city_dir, files in city_files.items():
        if files:
            # Sort by modification time, newest first
            newest = max(files, key=lambda f: f.stat().st_mtime)
            latest_files.append(newest)
    
    return latest_files


def extract_stats_from_excel(xlsx_path: Path) -> dict:
    """Extract stats from a city Excel file's Stats sheet."""
    stats = {
        "city": xlsx_path.parent.name,
        "file": xlsx_path.name,
        "leads_count": 0,
        "hotels_scraped": 0,
        "with_website": 0,
        "booking_found": 0,
        "tier1": 0,
        "tier2": 0,
        "with_phone": 0,
        "with_email": 0,
    }
    
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        
        # Get leads count from Leads sheet
        if "Leads" in wb.sheetnames:
            ws_leads = wb["Leads"]
            stats["leads_count"] = ws_leads.max_row - 1  # Subtract header
        
        # Try to parse Stats sheet for metrics
        if "Stats" in wb.sheetnames:
            ws_stats = wb["Stats"]
            
            # Read all cells into a dict for easier parsing
            cells = {}
            for row in ws_stats.iter_rows(max_row=30, max_col=6):
                for cell in row:
                    if cell.value:
                        cells[(cell.row, cell.column)] = str(cell.value)
            
            # Parse known positions (based on our dashboard layout)
            for (row, col), value in cells.items():
                val_lower = value.lower()
                
                # Look for our metric labels and get the value in next column
                next_val = cells.get((row, col + 1), "")
                
                # Extract number from value like "3,328" or "576 (33.0%)"
                def extract_num(s):
                    if not s:
                        return 0
                    # Remove commas and get first number
                    import re
                    match = re.search(r'[\d,]+', s.replace(',', ''))
                    if match:
                        try:
                            return int(match.group().replace(',', ''))
                        except:
                            pass
                    return 0
                
                if "hotels scraped" in val_lower:
                    stats["hotels_scraped"] = extract_num(next_val)
                elif "with website" in val_lower and "rate" not in val_lower:
                    stats["with_website"] = extract_num(next_val)
                elif "booking found" in val_lower:
                    stats["booking_found"] = extract_num(next_val)
                elif "tier 1" in val_lower:
                    stats["tier1"] = extract_num(next_val)
                elif "tier 2" in val_lower:
                    stats["tier2"] = extract_num(next_val)
                elif "with phone" in val_lower:
                    stats["with_phone"] = extract_num(next_val)
                elif "with email" in val_lower:
                    stats["with_email"] = extract_num(next_val)
        
        wb.close()
        
    except Exception as e:
        print(f"  Warning: Could not read {xlsx_path.name}: {e}")
    
    return stats


def normalize_country_name(name: str) -> str:
    """Normalize country name for display (e.g., usa -> USA)."""
    # Special cases
    if name.lower() == "usa":
        return "USA"
    return name.title()


def generate_country_stats(country_name: str, city_stats: list[dict]) -> Path:
    """Generate country-level aggregate stats Excel file."""
    
    # Find actual country path (case-insensitive)
    sadie_path = get_sadie_path()
    country_path = None
    for item in sadie_path.iterdir():
        if item.is_dir() and item.name.lower() == country_name.lower():
            country_path = item
            break
    
    if not country_path or not country_path.exists():
        print(f"Country folder not found: {country_name}")
        return None
    
    display_name = normalize_country_name(country_name)
    
    # Aggregate stats
    totals = {
        "hotels_scraped": sum(c["hotels_scraped"] for c in city_stats),
        "with_website": sum(c["with_website"] for c in city_stats),
        "booking_found": sum(c["booking_found"] for c in city_stats),
        "leads_count": sum(c["leads_count"] for c in city_stats),
        "tier1": sum(c["tier1"] for c in city_stats),
        "tier2": sum(c["tier2"] for c in city_stats),
        "with_phone": sum(c["with_phone"] for c in city_stats),
        "with_email": sum(c["with_email"] for c in city_stats),
    }
    
    # Calculate rates
    website_rate = round(totals["with_website"] / totals["hotels_scraped"] * 100, 1) if totals["hotels_scraped"] else 0
    detection_rate = round(totals["booking_found"] / totals["with_website"] * 100, 1) if totals["with_website"] else 0
    tier1_rate = round(totals["tier1"] / totals["booking_found"] * 100, 1) if totals["booking_found"] else 0
    
    # Create workbook
    wb = Workbook()
    
    # Styles
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="455A64", end_color="455A64", fill_type="solid")
    label_font = Font(size=10)
    value_font = Font(bold=True, size=12)
    green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='BDBDBD'),
        right=Side(style='thin', color='BDBDBD'),
        top=Side(style='thin', color='BDBDBD'),
        bottom=Side(style='thin', color='BDBDBD')
    )
    
    # =========================================================================
    # SHEET 1: AGGREGATE STATS
    # =========================================================================
    ws = wb.active
    ws.title = "Country Stats"
    
    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 18
    
    # Title
    ws.merge_cells('A1:E1')
    c = ws.cell(row=1, column=1, value=f"{display_name.upper()} — AGGREGATE STATS")
    c.font = title_font
    c.fill = title_fill
    c.alignment = center
    ws.row_dimensions[1].height = 35
    
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=2, column=1).font = Font(italic=True, size=9)
    
    # FUNNEL
    ws.merge_cells('A4:B4')
    c = ws.cell(row=4, column=1, value="AGGREGATE FUNNEL")
    c.font = section_font
    c.fill = section_fill
    c.alignment = center
    
    funnel_data = [
        ("Hotels Scraped", f"{totals['hotels_scraped']:,}"),
        ("With Website", f"{totals['with_website']:,} ({website_rate}%)"),
        ("Booking Found", f"{totals['booking_found']:,} ({detection_rate}%)"),
    ]
    for i, (label, val) in enumerate(funnel_data):
        r = 5 + i
        ws.cell(row=r, column=1, value=label).font = label_font
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2, value=val).font = value_font
        ws.cell(row=r, column=2).alignment = right
        ws.cell(row=r, column=2).border = thin_border
        if i == 2:
            ws.cell(row=r, column=1).fill = green_fill
            ws.cell(row=r, column=2).fill = green_fill
    
    # LEAD QUALITY
    ws.merge_cells('D4:E4')
    c = ws.cell(row=4, column=4, value="LEAD QUALITY")
    c.font = section_font
    c.fill = section_fill
    c.alignment = center
    
    quality_data = [
        ("Tier 1 (Known Engine)", f"{totals['tier1']:,} ({tier1_rate}%)"),
        ("Tier 2 (Unknown Engine)", f"{totals['tier2']:,} ({100-tier1_rate:.1f}%)"),
        ("Total Actionable", f"{totals['booking_found']:,}"),
    ]
    for i, (label, val) in enumerate(quality_data):
        r = 5 + i
        ws.cell(row=r, column=4, value=label).font = label_font
        ws.cell(row=r, column=4).border = thin_border
        ws.cell(row=r, column=5, value=val).font = value_font
        ws.cell(row=r, column=5).alignment = right
        ws.cell(row=r, column=5).border = thin_border
    
    # CITIES BREAKDOWN
    ws.merge_cells('A9:E9')
    c = ws.cell(row=9, column=1, value="CITIES BREAKDOWN")
    c.font = section_font
    c.fill = section_fill
    c.alignment = center
    
    # Headers
    headers = ["City", "Scraped", "Booking Found", "Tier 1", "Tier 2"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=10, column=col, value=h)
        c.font = Font(bold=True)
        c.border = thin_border
        c.alignment = center
    
    # City rows
    for i, city in enumerate(sorted(city_stats, key=lambda x: x["booking_found"], reverse=True)):
        r = 11 + i
        ws.cell(row=r, column=1, value=city["city"]).border = thin_border
        ws.cell(row=r, column=2, value=city["hotels_scraped"]).border = thin_border
        ws.cell(row=r, column=2).alignment = right
        ws.cell(row=r, column=3, value=city["booking_found"]).border = thin_border
        ws.cell(row=r, column=3).alignment = right
        ws.cell(row=r, column=4, value=city["tier1"]).border = thin_border
        ws.cell(row=r, column=4).alignment = right
        ws.cell(row=r, column=5, value=city["tier2"]).border = thin_border
        ws.cell(row=r, column=5).alignment = right
    
    # Save - single file, overwrites each time
    # Keep USA uppercase, title case others
    display_name = "USA" if country_name.lower() == "usa" else country_name.title()
    output_path = country_path / f"{display_name} Stats.xlsx"
    wb.save(output_path)
    
    return output_path


def process_country(country_name: str):
    """Process a single country and generate stats."""
    country_path = get_sadie_path() / country_name.title()
    
    if not country_path.exists():
        print(f"❌ Country folder not found: {country_path}")
        return
    
    print(f"\n📊 Processing {country_name.title()}...")
    
    # Find all city files
    city_files = find_city_files(country_path)
    
    if not city_files:
        print(f"  No city data files found in {country_path}")
        return
    
    print(f"  Found {len(city_files)} city files")
    
    # Extract stats from each
    city_stats = []
    for xlsx in city_files:
        stats = extract_stats_from_excel(xlsx)
        city_stats.append(stats)
        print(f"    ✓ {stats['city']}: {stats['booking_found']} booking found")
    
    # Generate country stats
    output = generate_country_stats(country_name, city_stats)
    
    if output:
        print(f"\n✅ Saved: {output}")
        
        # Print summary
        total_scraped = sum(c["hotels_scraped"] for c in city_stats)
        total_booking = sum(c["booking_found"] for c in city_stats)
        print(f"   Total scraped: {total_scraped:,}")
        print(f"   Total booking found: {total_booking:,}")


def main():
    parser = argparse.ArgumentParser(description="Generate country-level aggregate stats")
    parser.add_argument("--country", help="Process specific country (e.g., 'usa', 'australia')")
    args = parser.parse_args()
    
    sadie_path = get_sadie_path()
    if not sadie_path.exists():
        print(f"❌ Sadie Lead Gen folder not found: {sadie_path}")
        sys.exit(1)
    
    print(f"📁 Scanning: {sadie_path}")
    
    if args.country:
        # Process specific country
        process_country(args.country)
    else:
        # Process all countries
        for item in sadie_path.iterdir():
            if item.is_dir() and not item.name.startswith("."):
                process_country(item.name)


if __name__ == "__main__":
    main()

