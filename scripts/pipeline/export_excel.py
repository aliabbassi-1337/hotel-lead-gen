#!/usr/bin/env python3
"""
Sadie Excel Export - Export Leads with Stats Sheet
===================================================
Creates an Excel file with leads data and a separate stats sheet.

Usage:
    python3 sadie_excel_export.py --input detector_output/sydney_leads_post.csv --city sydney
    python3 sadie_excel_export.py --input detector_output/ocean_city_leads_post.csv --city ocean_city --scraper scraper_output/ocean_city_hotels.csv
"""

import csv
import sys
import os
import argparse
from datetime import datetime
from collections import Counter
from urllib.parse import urlparse

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Install with: pip3 install openpyxl")
    sys.exit(1)


def load_csv(filepath: str) -> tuple:
    """Load CSV file and return headers and rows."""
    if not filepath or not os.path.exists(filepath):
        return [], []
    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = reader.fieldnames if reader.fieldnames else []
    return headers, rows


def extract_domain(url: str) -> str:
    """Extract domain from URL."""
    try:
        return urlparse(url).netloc.lower().replace("www.", "")
    except:
        return ""


def generate_stats(city: str, leads_rows: list, scraper_rows: list = None, detector_rows: list = None) -> list:
    """Generate stats rows for the stats sheet."""
    
    stats = []
    stats.append(["LEAD GENERATION FUNNEL", ""])
    stats.append(["City", city.replace("_", " ").title()])
    stats.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    stats.append(["", ""])
    
    # Calculate detector stats first (use detector_rows if provided)
    if detector_rows:
        detector_total = len(detector_rows)
        detector_with_booking = sum(1 for r in detector_rows if r.get("booking_url", "").strip())
        detector_with_website = sum(1 for r in detector_rows if r.get("website", "").strip())
    else:
        detector_total = 0
        detector_with_booking = 0
        detector_with_website = 0
    
    # Calculate scraper stats - BUT if detector is smaller (split file), use detector as base
    scraper_total = len(scraper_rows) if scraper_rows else 0
    scraper_with_website = sum(1 for r in scraper_rows if r.get("website", "").strip()) if scraper_rows else 0
    
    # For split files: detector has fewer rows than scraper, use detector as the base
    if detector_total > 0 and (scraper_total == 0 or detector_total < scraper_total * 0.5):
        # This is a split file - use detector totals as the "scraped" base
        scraper_total = detector_total
        scraper_with_website = detector_with_website
    
    # Final leads stats
    leads_total = len(leads_rows)
    leads_with_booking = sum(1 for r in leads_rows if r.get("booking_url", "").strip())
    
    known_engine_values = ["unknown", "unknown_third_party", "unknown_booking_api", 
                           "proprietary_or_same_domain", "contact_only", ""]
    leads_with_known_engine = sum(1 for r in leads_rows 
        if r.get("booking_engine", "").strip() and r.get("booking_engine") not in known_engine_values)
    
    # =========================================================================
    # FUNNEL VISUALIZATION
    # =========================================================================
    stats.append(["=== FUNNEL OVERVIEW ===", ""])
    stats.append(["", ""])
    
    if scraper_total > 0:
        stats.append(["Hotels Scraped (Google Maps)", scraper_total])
        stats.append(["Hotels with Website", f"{scraper_with_website} ({round(scraper_with_website/scraper_total*100, 1)}%)"])
    
    if detector_total > 0:
        stats.append(["Websites Processed", detector_total])
        detection_rate = round(detector_with_booking/detector_total*100, 1) if detector_total > 0 else 0
        stats.append(["Booking URL Detected", f"{detector_with_booking} ({detection_rate}%) ← DETECTION RATE"])
    
    stats.append(["Final Clean Leads", leads_total])
    stats.append(["Leads with Booking URL", f"{leads_with_booking} ({round(leads_with_booking/leads_total*100, 1)}%)"])
    stats.append(["Leads with Known Engine", f"{leads_with_known_engine} ({round(leads_with_known_engine/leads_total*100, 1)}%)"])
    
    stats.append(["", ""])
    
    # =========================================================================
    # KEY METRICS
    # =========================================================================
    stats.append(["=== KEY METRICS ===", ""])
    
    if scraper_total > 0:
        stats.append(["Website Coverage Rate", f"{round(scraper_with_website/scraper_total*100, 1)}%"])
    
    if detector_total > 0:
        stats.append(["Booking Detection Rate", f"{round(detector_with_booking/detector_total*100, 1)}%"])
    
    if scraper_with_website > 0:
        stats.append(["End-to-End Conversion (Website → Booking)", f"{round(leads_with_booking/scraper_with_website*100, 1)}%"])
    
    if scraper_total > 0:
        stats.append(["Overall Funnel Conversion", f"{round(leads_with_booking/scraper_total*100, 1)}%"])
    
    stats.append(["", ""])
    
    # =========================================================================
    # LEAD QUALITY TIERS
    # =========================================================================
    tier1 = sum(1 for r in leads_rows 
        if r.get("booking_url", "").strip() and 
        r.get("booking_engine", "").strip() and 
        r.get("booking_engine") not in known_engine_values)
    tier2 = sum(1 for r in leads_rows 
        if r.get("booking_url", "").strip() and 
        (not r.get("booking_engine", "").strip() or r.get("booking_engine") in known_engine_values))
    tier3 = leads_total - tier1 - tier2
    
    stats.append(["=== LEAD QUALITY ===", ""])
    stats.append(["Tier 1 (Booking URL + Known Engine)", f"{tier1} ({round(tier1/leads_total*100, 1)}%)"])
    stats.append(["Tier 2 (Booking URL + Unknown Engine)", f"{tier2} ({round(tier2/leads_total*100, 1)}%)"])
    stats.append(["Tier 3 (No Booking URL)", f"{tier3} ({round(tier3/leads_total*100, 1)}%)"])
    stats.append(["", ""])
    stats.append(["ACTIONABLE LEADS (Tier 1 + 2)", tier1 + tier2])
    stats.append(["", ""])
    
    # =========================================================================
    # CONTACT INFO
    # =========================================================================
    with_phone = sum(1 for r in leads_rows 
        if r.get("phone_google", "").strip() or r.get("phone_website", "").strip() or r.get("Phone Number", "").strip())
    with_email = sum(1 for r in leads_rows 
        if r.get("email", "").strip() or r.get("Email", "").strip())
    
    stats.append(["=== CONTACT INFO ===", ""])
    stats.append(["With Phone", f"{with_phone} ({round(with_phone/leads_total*100, 1)}%)"])
    stats.append(["With Email", f"{with_email} ({round(with_email/leads_total*100, 1)}%)"])
    stats.append(["", ""])
    
    # =========================================================================
    # ENGINE BREAKDOWN
    # =========================================================================
    engine_counts = Counter(r.get("booking_engine", "") or r.get("PMS", "") 
        for r in leads_rows if (r.get("booking_engine", "") or r.get("PMS", "")).strip())
    
    if engine_counts:
        stats.append(["=== TOP BOOKING ENGINES ===", ""])
        for engine, count in engine_counts.most_common(15):
            stats.append([engine, count])
    
    return stats


def create_excel_export(input_file: str, city: str, scraper_file: str = None, 
                        detector_file: str = None, output_file: str = None, hubspot_format: bool = False):
    """Create Excel file with leads and stats sheets."""
    
    # Load data
    headers, leads_rows = load_csv(input_file)
    _, scraper_rows = load_csv(scraper_file) if scraper_file else ([], [])
    _, detector_rows = load_csv(detector_file) if detector_file else ([], [])
    
    if not leads_rows:
        print(f"Error: No data in {input_file}")
        return
    
    # Filter out debug columns - these are NOT for sales people
    debug_columns = [
        "error", "detection_method", "screenshot_path", 
        "booking_engine_domain", "latitude", "longitude"
    ]
    headers = [h for h in headers if h not in debug_columns]
    
    # Create workbook
    wb = Workbook()
    
    # =========================================================================
    # SHEET 1: LEADS
    # =========================================================================
    ws_leads = wb.active
    ws_leads.title = "Leads"
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws_leads.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data with center alignment
    center_align = Alignment(horizontal="center", vertical="center")
    for row_idx, row in enumerate(leads_rows, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws_leads.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
            cell.alignment = center_align
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for row in range(1, min(len(leads_rows) + 2, 100)):  # Check first 100 rows
            cell_value = ws_leads.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws_leads.column_dimensions[column].width = adjusted_width
    
    # Freeze header row
    ws_leads.freeze_panes = "A2"
    
    # =========================================================================
    # SHEET 2: STATS (Dashboard Layout - 2 columns per row)
    # =========================================================================
    ws_stats = wb.create_sheet(title="Stats")
    
    # Style definitions
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="455A64", end_color="455A64", fill_type="solid")
    label_font = Font(size=10)
    value_font = Font(bold=True, size=12)
    green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='BDBDBD'),
        right=Side(style='thin', color='BDBDBD'),
        top=Side(style='thin', color='BDBDBD'),
        bottom=Side(style='thin', color='BDBDBD')
    )
    
    # Calculate all stats
    detector_total = len(detector_rows) if detector_rows else 0
    detector_with_booking = sum(1 for r in detector_rows if r.get("booking_url", "").strip()) if detector_rows else 0
    detector_with_website = sum(1 for r in detector_rows if r.get("website", "").strip()) if detector_rows else 0
    
    # Scraper stats - always use actual scraper file if provided
    scraper_total = len(scraper_rows) if scraper_rows else detector_total
    scraper_with_website = sum(1 for r in scraper_rows if r.get("website", "").strip()) if scraper_rows else detector_with_website
    
    leads_total = len(leads_rows)
    leads_with_booking = sum(1 for r in leads_rows if r.get("booking_url", "").strip())
    booking_found = leads_with_booking  # Use final post-processed count
    
    known_engine_values = ["unknown", "unknown_third_party", "unknown_booking_api", 
                           "proprietary_or_same_domain", "contact_only", ""]
    
    tier1 = sum(1 for r in leads_rows 
        if r.get("booking_url", "").strip() and 
        r.get("booking_engine", "").strip() and 
        r.get("booking_engine") not in known_engine_values)
    tier2 = leads_with_booking - tier1
    tier3 = leads_total - tier1 - tier2
    actionable = tier1 + tier2
    
    with_phone = sum(1 for r in leads_rows 
        if r.get("phone_google", "").strip() or r.get("phone_website", "").strip() or r.get("Phone Number", "").strip())
    with_email = sum(1 for r in leads_rows 
        if r.get("email", "").strip() or r.get("Email", "").strip())
    
    # Set column widths (2 sections: A-C and E-G with D as spacer)
    widths = {"A": 22, "B": 15, "C": 2, "D": 22, "E": 15}
    for col, w in widths.items():
        ws_stats.column_dimensions[col].width = w
    
    # ROW 1: Title
    ws_stats.merge_cells('A1:E1')
    c = ws_stats.cell(row=1, column=1, value=f"LEAD GENERATION DASHBOARD — {city.replace('_', ' ').upper()}")
    c.font = title_font
    c.fill = title_fill
    c.alignment = center
    ws_stats.row_dimensions[1].height = 35
    
    # =========================================================================
    # ROW 3-6: FUNNEL (count + percentage on same row)
    # =========================================================================
    
    # Calculate rates
    website_rate = round(scraper_with_website/scraper_total*100, 1) if scraper_total else 0
    detection_rate = round(leads_with_booking/scraper_with_website*100, 1) if scraper_with_website else 0
    overall_rate = round(leads_with_booking/scraper_total*100, 1) if scraper_total else 0
    
    ws_stats.merge_cells('A3:B3')
    c = ws_stats.cell(row=3, column=1, value="FUNNEL")
    c.font = section_font
    c.fill = section_fill
    c.alignment = center
    
    funnel_data = [
        ("Hotels Scraped", f"{scraper_total:,}" if scraper_total else "N/A"),
        ("With Website", f"{scraper_with_website:,} ({website_rate}%)" if scraper_total else "N/A"),
        ("Booking Found", f"{leads_with_booking:,} ({detection_rate}%)"),
    ]
    for i, (label, val) in enumerate(funnel_data):
        r = 4 + i
        ws_stats.cell(row=r, column=1, value=label).font = label_font
        ws_stats.cell(row=r, column=1).border = thin_border
        ws_stats.cell(row=r, column=2, value=val).font = value_font
        ws_stats.cell(row=r, column=2).alignment = right
        ws_stats.cell(row=r, column=2).border = thin_border
        if i == 2:  # Highlight booking found
            ws_stats.cell(row=r, column=1).fill = green_fill
            ws_stats.cell(row=r, column=2).fill = green_fill
    
    # Right side of row 3-6: LEAD QUALITY (only for leads WITH booking URLs)
    ws_stats.merge_cells('D3:E3')
    c = ws_stats.cell(row=3, column=4, value="LEAD QUALITY (of Booking Found)")
    c.font = section_font
    c.fill = section_fill
    c.alignment = center
    
    # Percentages based on leads_with_booking, not total leads
    tier1_pct = round(tier1/leads_with_booking*100, 1) if leads_with_booking else 0
    tier2_pct = round(tier2/leads_with_booking*100, 1) if leads_with_booking else 0
    
    quality_data = [
        ("Tier 1 (Known Engine)", f"{tier1:,} ({tier1_pct}%)"),
        ("Tier 2 (Unknown Engine)", f"{tier2:,} ({tier2_pct}%)"),
        ("Total", f"{leads_with_booking:,} (100%)"),
    ]
    for i, (label, val) in enumerate(quality_data):
        r = 4 + i
        ws_stats.cell(row=r, column=4, value=label).font = label_font
        ws_stats.cell(row=r, column=4).border = thin_border
        ws_stats.cell(row=r, column=5, value=val).font = value_font
        ws_stats.cell(row=r, column=5).alignment = right
        ws_stats.cell(row=r, column=5).border = thin_border
    
    # =========================================================================
    # ROW 8-10: CONTACT INFO (left) | TOP ENGINES preview (right)
    # =========================================================================
    
    ws_stats.merge_cells('A8:B8')
    c = ws_stats.cell(row=8, column=1, value="CONTACT INFO")
    c.font = section_font
    c.fill = section_fill
    c.alignment = center
    
    contact_data = [
        ("With Phone", f"{with_phone:,} ({round(with_phone/leads_total*100, 1)}%)"),
        ("With Email", f"{with_email:,} ({round(with_email/leads_total*100, 1)}%)"),
    ]
    for i, (label, val) in enumerate(contact_data):
        r = 9 + i
        ws_stats.cell(row=r, column=1, value=label).font = label_font
        ws_stats.cell(row=r, column=1).border = thin_border
        ws_stats.cell(row=r, column=2, value=val).font = value_font
        ws_stats.cell(row=r, column=2).alignment = right
        ws_stats.cell(row=r, column=2).border = thin_border
    
    # =========================================================================
    # ROW 8+: BOOKING ENGINES (right side, next to contact info)
    # =========================================================================
    
    ws_stats.merge_cells('D8:E8')
    c = ws_stats.cell(row=8, column=4, value="TOP ENGINES")
    c.font = section_font
    c.fill = section_fill
    c.alignment = center
    
    engine_counts = Counter(r.get("booking_engine", "") or r.get("PMS", "") 
        for r in leads_rows if (r.get("booking_engine", "") or r.get("PMS", "")).strip())
    engines = engine_counts.most_common(10)
    
    for i, (engine, count) in enumerate(engines):
        r = 9 + i
        ws_stats.cell(row=r, column=4, value=engine).font = label_font
        ws_stats.cell(row=r, column=4).border = thin_border
        ws_stats.cell(row=r, column=5, value=count).font = value_font
        ws_stats.cell(row=r, column=5).alignment = right
        ws_stats.cell(row=r, column=5).border = thin_border
    
    # Freeze title row
    ws_stats.freeze_panes = "A2"
    
    # =========================================================================
    # SAVE FILE
    # =========================================================================
    if not output_file:
        base_name = os.path.splitext(os.path.basename(input_file))[0]
        output_file = f"{base_name}.xlsx"
    
    wb.save(output_file)
    print(f"✅ Excel file saved: {output_file}")
    print(f"   - Sheet 1 'Leads': {len(leads_rows):,} rows")
    print(f"   - Sheet 2 'Stats': Dashboard with funnel metrics")


def main():
    parser = argparse.ArgumentParser(description="Export leads to Excel with stats sheet")
    parser.add_argument("--input", required=True, help="Input CSV file (post-processed leads)")
    parser.add_argument("--city", required=True, help="City name for stats")
    parser.add_argument("--scraper", help="Scraper output CSV (for full funnel stats)")
    parser.add_argument("--detector", help="Detector output CSV (for detection rate)")
    parser.add_argument("--output", help="Output Excel file path")
    
    args = parser.parse_args()
    
    create_excel_export(
        input_file=args.input,
        city=args.city,
        scraper_file=args.scraper,
        detector_file=args.detector,
        output_file=args.output
    )


if __name__ == "__main__":
    main()

